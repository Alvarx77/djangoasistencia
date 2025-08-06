# alumnos/views.py
import pandas as pd
import unicodedata
from django.shortcuts import render, redirect
from .models import Alumno, Curso, AsistenciaMensual, DiasClaseMensual
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.views.decorators.csrf import csrf_protect
from datetime import datetime
from django.urls import reverse
from django.http import JsonResponse
from django.db.models import Sum
from django.views.decorators.http import require_GET


import io
from collections import defaultdict
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, NamedStyle
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter


def normalizar(texto):
    if pd.isna(texto):
        return ""
    texto = str(texto).strip().upper()
    texto = ''.join(c for c in unicodedata.normalize('NFKD', texto) if not unicodedata.combining(c))
    return texto

@csrf_protect
def cargar_excel(request):
    if request.method == 'POST':
        if 'eliminar_bd' in request.POST:
            Alumno.objects.all().delete()
            Curso.objects.all().delete()
            messages.success(request, "✅ Se eliminó toda la base de datos correctamente.")
            return redirect('cargar_excel')

        archivo = request.FILES.get('excel_file')
        if archivo:
            df = pd.read_excel(archivo)
            df = df[df['Fecha Retiro'] == pd.to_datetime('1900-01-01')]

            for _, fila in df.iterrows():
                nombre = normalizar(fila['Nombres'])
                apellido_paterno = normalizar(fila['Apellido Paterno'])
                apellido_materno = normalizar(fila['Apellido Materno'])
                nombre_completo = f"{apellido_paterno} {apellido_materno} {nombre}".strip()

                desc_grado = normalizar(fila['Desc Grado'])
                letra_curso = normalizar(fila['Letra Curso'])
                nombre_curso = f"{desc_grado} {letra_curso}".strip()

                curso, _ = Curso.objects.get_or_create(nombre=nombre_curso)
                Alumno.objects.get_or_create(nombre_completo=nombre_completo, curso=curso)

            messages.success(request, "✅ Archivo Excel cargado exitosamente.")
            return redirect('lista_alumnos')

    return render(request, 'alumnos/cargar_excel.html')


def _to_upper_clean(s: str) -> str:
    return " ".join((s or "").strip().upper().split())

def _build_nombre_completo(ap_paterno: str, ap_materno: str, nombres: str) -> str:
    ap_paterno = _to_upper_clean(ap_paterno)
    ap_materno = _to_upper_clean(ap_materno)
    nombres = _to_upper_clean(nombres)
    return " ".join([p for p in [ap_paterno, ap_materno, nombres] if p])

@login_required
def lista_alumnos(request):
    cursos = Curso.objects.all().order_by('nombre')

    # Filtros (por nombre de curso, como ya lo tenías)
    curso_filtrado = request.GET.get('curso', '').strip()
    if not curso_filtrado and cursos.exists():
        curso_filtrado = cursos.first().nombre

    nombre_filtrado = (request.GET.get('nombre') or '').strip().upper()

    # Acciones (todo POST se maneja aquí)
    if request.method == 'POST':
        action = request.POST.get('action')

        # Para volver con filtros actuales tras la acción
        q_params = f"?curso={curso_filtrado}&nombre={nombre_filtrado}".replace(' ', '%20')

        if action == 'add':
            curso_id = request.POST.get('curso_id')
            ap_paterno = request.POST.get('ap_paterno', '')
            ap_materno = request.POST.get('ap_materno', '')
            nombres = request.POST.get('nombres', '')
            nombre_completo = _build_nombre_completo(ap_paterno, ap_materno, nombres)

            if not curso_id or not nombre_completo:
                messages.error(request, "Completa todos los campos para agregar el estudiante.")
                return redirect(reverse('lista_alumnos') + q_params)

            try:
                curso = Curso.objects.get(id=curso_id)
            except Curso.DoesNotExist:
                messages.error(request, "El curso seleccionado no existe.")
                return redirect(reverse('lista_alumnos') + q_params)

            Alumno.objects.create(nombre_completo=nombre_completo, curso=curso)
            messages.success(request, f"Alumno agregado: {nombre_completo}")
            return redirect(reverse('lista_alumnos') + q_params)

        if action == 'edit':
            alumno_id = request.POST.get('alumno_id')
            curso_id = request.POST.get('curso_id')
            ap_paterno = request.POST.get('ap_paterno', '')
            ap_materno = request.POST.get('ap_materno', '')
            nombres = request.POST.get('nombres', '')
            nombre_completo = _build_nombre_completo(ap_paterno, ap_materno, nombres)

            if not alumno_id or not curso_id or not nombre_completo:
                messages.error(request, "Completa todos los campos para editar el estudiante.")
                return redirect(reverse('lista_alumnos') + q_params)

            try:
                alumno = Alumno.objects.get(id=alumno_id)
                curso = Curso.objects.get(id=curso_id)
            except (Alumno.DoesNotExist, Curso.DoesNotExist):
                messages.error(request, "Alumno o curso no encontrado.")
                return redirect(reverse('lista_alumnos') + q_params)

            alumno.nombre_completo = nombre_completo
            alumno.curso = curso
            alumno.save()
            messages.success(request, f"Alumno actualizado: {nombre_completo}")
            return redirect(reverse('lista_alumnos') + q_params)

        if action == 'delete':
            alumno_id = request.POST.get('alumno_id')
            try:
                alumno = Alumno.objects.get(id=alumno_id)
                nombre = alumno.nombre_completo
                alumno.delete()
                messages.success(request, f"Alumno eliminado: {nombre}")
            except Alumno.DoesNotExist:
                messages.error(request, "Alumno no encontrado.")
            return redirect(reverse('lista_alumnos') + q_params)

        # Acción desconocida
        messages.error(request, "Acción no válida.")
        return redirect(reverse('lista_alumnos') + q_params)

    # Consulta base
    alumnos = Alumno.objects.all()
    if curso_filtrado:
        alumnos = alumnos.filter(curso__nombre=curso_filtrado)
    if nombre_filtrado:
        alumnos = alumnos.filter(nombre_completo__icontains=nombre_filtrado)

    alumnos = alumnos.select_related('curso').order_by('nombre_completo')

    return render(request, 'alumnos/lista_alumnos.html', {
        'alumnos': alumnos,
        'cursos': cursos,
        'curso_filtrado': curso_filtrado,
        'nombre_filtrado': nombre_filtrado,
    })


# Alias para mantener la ruta y los templates que usan {% url 'dashboard' %}
@login_required
def dashboard(request):
    return dashboard(request)



@login_required
def dashboard(request):
    cursos = Curso.objects.all().order_by('nombre')
    mes_str = request.GET.get('mes') or datetime.today().strftime('%Y-%m')
    mes_date = datetime.strptime(mes_str, '%Y-%m').date().replace(day=1)
    umbral_critico = 85  # % mínimo deseado

    resumen_cursos = []

    for curso in cursos:
        alumnos = Alumno.objects.filter(curso=curso)
        total_porcentaje = 0
        total_validos = 0

        for alumno in alumnos:
            asistencia = AsistenciaMensual.objects.filter(alumno=alumno, mes=mes_date).first()
            dias = DiasClaseMensual.objects.filter(curso=curso, mes=mes_date).first()
            dias_total = dias.dias_clases if dias else 0
            if asistencia and dias_total:
                porcentaje = round((asistencia.presentes / dias_total) * 100, 1)
                total_porcentaje += porcentaje
                total_validos += 1

        promedio = round(total_porcentaje / total_validos, 1) if total_validos else 0
        resumen_cursos.append({
            'curso': curso,
            'promedio': promedio,
        })

    mejores_cursos = sorted(resumen_cursos, key=lambda x: x['promedio'], reverse=True)[:3]
    cursos_criticos = [rc for rc in resumen_cursos if rc['promedio'] < umbral_critico]

    return render(request, 'alumnos/dashboard.html', {
        'mejores_cursos': mejores_cursos,
        'cursos_criticos': cursos_criticos,
        'mes_actual': mes_str,
        'umbral_critico': umbral_critico,
    })



# ===========================
# Vista principal: Asistencia Mensual
# ===========================
@login_required
def asistencia_mensual(request):
    cursos = Curso.objects.all().order_by('nombre')
    curso_id = request.GET.get('curso') or (cursos.first().id if cursos.exists() else None)
    mes_str = request.GET.get('mes') or datetime.today().strftime('%Y-%m')
    mes_date = datetime.strptime(mes_str, '%Y-%m').date().replace(day=1)

    curso = Curso.objects.filter(id=curso_id).first()
    alumnos = Alumno.objects.filter(curso=curso).order_by('nombre_completo') if curso else []

    if request.method == 'POST' and curso:
        # Guarda SIEMPRE (aunque no haya días definidos)
        for alumno in alumnos:
            presentes = int(request.POST.get(f'presentes_{alumno.id}', 0) or 0)
            inasistentes = int(request.POST.get(f'inasistentes_{alumno.id}', 0) or 0)

            asistencia, created = AsistenciaMensual.objects.get_or_create(
                alumno=alumno,
                mes=mes_date,
                defaults={'curso': curso}  # ✅ evita NOT NULL en el INSERT
            )
            # En caso de registros antiguos sin curso:
            if not asistencia.curso_id:
                asistencia.curso = curso

            asistencia.presentes = presentes
            asistencia.inasistentes = inasistentes
            asistencia.save()

        # Guarda días de clases del curso en el mes (aunque sea 0)
        dias = int(request.POST.get('dias_clases', 0) or 0)
        dcm, _ = DiasClaseMensual.objects.get_or_create(curso=curso, mes=mes_date)
        dcm.dias_clases = dias
        dcm.save()

        messages.success(request, '✅ Asistencia actualizada correctamente.')
        return redirect(f'/asistencia_mensual/?curso={curso_id}&mes={mes_str}')

    # Carga para mostrar
    asistencias_dict = {
        a.alumno.id: a
        for a in AsistenciaMensual.objects.filter(mes=mes_date, alumno__curso=curso)
    }
    dcm = DiasClaseMensual.objects.filter(curso=curso, mes=mes_date).first()
    dias_total = dcm.dias_clases if dcm else 0

    alumnos_asistencia = []
    total_porcentajes = 0
    for alumno in alumnos:
        asistencia = asistencias_dict.get(alumno.id)
        if asistencia and dias_total > 0:
            porcentaje = round((asistencia.presentes / dias_total) * 100, 1)
        else:
            porcentaje = 0.0
        total_porcentajes += porcentaje
        alumnos_asistencia.append({
            'alumno': alumno,
            'asistencia': asistencia,
            'porcentaje': porcentaje,
        })

    promedio_asistencia = round(total_porcentajes / len(alumnos_asistencia), 1) if alumnos_asistencia else 0.0

    return render(request, 'alumnos/asistencia_mensual.html', {
        'cursos': cursos,
        'curso': curso,
        'curso_id': curso_id,
        'mes': mes_str,
        'dias_clases': dias_total,
        'alumnos_asistencia': alumnos_asistencia,
        'promedio_asistencia': promedio_asistencia,
        'total_alumnos': len(alumnos_asistencia),
    })



# ===========================
# AJAX: Auto-guardado de presentes/inasistentes por alumno
# ===========================
@login_required
def ajax_actualizar_asistencia(request):
    """
    Autosave: presentes/inasistentes de un alumno.
    POST: alumno_id, curso_id, mes(YYYY-MM), presentes, inasistentes
    Resp: {ok, porcentaje}
    """
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'Método no permitido'}, status=405)

    try:
        alumno_id = int(request.POST.get('alumno_id'))
        curso_id = int(request.POST.get('curso_id'))
        mes_str = request.POST.get('mes')
        presentes = int(request.POST.get('presentes', 0) or 0)
        inasistentes = int(request.POST.get('inasistentes', 0) or 0)

        mes_date = datetime.strptime(mes_str, '%Y-%m').date().replace(day=1)
        alumno = Alumno.objects.get(id=alumno_id, curso_id=curso_id)
        curso = Curso.objects.get(id=curso_id)

        asistencia, _ = AsistenciaMensual.objects.update_or_create(
            alumno=alumno,
            mes=mes_date,
            defaults={
                'presentes': presentes,
                'inasistentes': inasistentes,
                'curso': curso,  # ✅ importante para NOT NULL
            }
        )

        dcm = DiasClaseMensual.objects.filter(curso=curso, mes=mes_date).first()
        dias_total = dcm.dias_clases if dcm else 0
        porcentaje = round((presentes / dias_total) * 100, 1) if dias_total else 0.0

        return JsonResponse({'ok': True, 'porcentaje': porcentaje})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=400)



# ===========================
# AJAX: Auto-guardado de días de clases del curso
# ===========================
@login_required
def ajax_actualizar_dias_clases(request):
    """
    Autosave: días de clases del curso en el mes.
    POST: curso_id, mes(YYYY-MM), dias_clases
    Resp: {ok}
    """
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'Método no permitido'}, status=405)

    try:
        curso_id = int(request.POST.get('curso_id'))
        mes_str = request.POST.get('mes')
        dias = int(request.POST.get('dias_clases', 0) or 0)

        mes_date = datetime.strptime(mes_str, '%Y-%m').date().replace(day=1)
        curso = Curso.objects.get(id=curso_id)

        dcm, _ = DiasClaseMensual.objects.get_or_create(curso=curso, mes=mes_date)
        dcm.dias_clases = dias
        dcm.save()

        return JsonResponse({'ok': True})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=400)

    
    
@login_required
def estadisticas(request):
    """
    Página principal de estadísticas: selector de mes + contenedores de podio/gráfico.
    Los datos se cargan por AJAX desde `ajax_estadisticas_mes`.
    """
    mes_str = request.GET.get('mes') or datetime.today().strftime('%Y-%m')
    return render(request, 'alumnos/estadisticas.html', {'mes': mes_str})


@login_required
@require_GET
def ajax_estadisticas_mes(request):
    """
    Devuelve % asistencia por curso para el mes dado.
    GET: mes=YYYY-MM
    Resp JSON:
    {
      ok: true,
      cursos: [
        {"curso_id":..., "curso":"...", "porcentaje": 0.0-100.0, "alumnos":N, "presentes_total":X, "dias_clases":D},
        ...
      ],
      top3: [ ... ] // top 3 por porcentaje desc
    }
    """
    mes_str = request.GET.get('mes')
    try:
        mes_date = datetime.strptime(mes_str, '%Y-%m').date().replace(day=1)
    except Exception:
        return JsonResponse({'ok': False, 'error': 'Parámetro mes inválido (YYYY-MM).'}, status=400)

    cursos = Curso.objects.all().order_by('nombre')

    # Mapa de días por curso en el mes
    dias_map = {
        d.curso_id: (d.dias_clases or 0)
        for d in DiasClaseMensual.objects.filter(mes=mes_date)
    }

    # Suma de presentes por curso (en el mes)
    presentes_por_curso = (
        AsistenciaMensual.objects
        .filter(mes=mes_date)
        .values('curso_id')
        .annotate(presentes_total=Sum('presentes'))
    )
    presentes_map = {x['curso_id']: (x['presentes_total'] or 0) for x in presentes_por_curso}

    data_cursos = []
    for c in cursos:
        n_alumnos = Alumno.objects.filter(curso=c).count()
        dias = int(dias_map.get(c.id, 0) or 0)  # si no hay registro, 0
        presentes_total = int(presentes_map.get(c.id, 0) or 0)

        # Evita divisiones por cero y NaN/Infinity
        if dias > 0 and n_alumnos > 0:
            denom = dias * n_alumnos
            porcentaje = (presentes_total / denom) * 100.0
        else:
            porcentaje = 0.0

        # Sanitiza (por si acaso)
        try:
            porcentaje = float(porcentaje)
        except Exception:
            porcentaje = 0.0
        if porcentaje != porcentaje or porcentaje == float('inf') or porcentaje == float('-inf'):
            porcentaje = 0.0

        # Recorta a 1 decimal y acota 0..100
        porcentaje = round(max(0.0, min(100.0, porcentaje)), 1)

        data_cursos.append({
            'curso_id': c.id,
            'curso': c.nombre,
            'porcentaje': porcentaje,
            'alumnos': n_alumnos,
            'presentes_total': presentes_total,
            'dias_clases': dias,
        })

    top3 = sorted(data_cursos, key=lambda x: x['porcentaje'], reverse=True)[:3]

    return JsonResponse({'ok': True, 'cursos': data_cursos, 'top3': top3})



@login_required
def reporte_cursos_mes(request):
    """
    Muestra, por mes y por curso, dos tablas:
    - Alumnos con 100% asistencia
    - Alumnos críticos (<85%)
    Depende de que exista DiasClaseMensual para el curso en el mes.
    """
    mes_str = request.GET.get('mes') or datetime.today().strftime('%Y-%m')
    try:
        mes_date = datetime.strptime(mes_str, '%Y-%m').date().replace(day=1)
    except Exception:
        mes_date = datetime.today().replace(day=1)
        mes_str = mes_date.strftime('%Y-%m')

    cursos = Curso.objects.all().order_by('nombre')

    dias_map = {
        d.curso_id: int(d.dias_clases or 0)
        for d in DiasClaseMensual.objects.filter(mes=mes_date)
    }

    asistencias_mes = (
        AsistenciaMensual.objects
        .filter(mes=mes_date)
        .select_related('alumno', 'curso')
    )
    asistencia_por_alumno = {a.alumno_id: a for a in asistencias_mes}

    cursos_data = []

    for c in cursos:
        dias_curso = dias_map.get(c.id, 0)
        alumnos = list(Alumno.objects.filter(curso=c).order_by('nombre_completo'))

        perfectos = []
        criticos = []

        if dias_curso and dias_curso > 0:
            for al in alumnos:
                a = asistencia_por_alumno.get(al.id)
                presentes = int(getattr(a, 'presentes', 0) or 0)
                inasistentes = int(getattr(a, 'inasistentes', 0) or 0)

                # Limites sanos
                if presentes < 0: presentes = 0
                if inasistentes < 0: inasistentes = 0
                if presentes > dias_curso: presentes = dias_curso
                if inasistentes > dias_curso: inasistentes = dias_curso

                pct = (presentes / dias_curso) * 100.0 if dias_curso > 0 else 0.0
                pct = round(max(0.0, min(100.0, pct)), 1)

                if presentes == dias_curso:
                    perfectos.append((al, presentes, inasistentes, pct))
                elif pct < 85.0:
                    criticos.append((al, presentes, inasistentes, pct))

        cursos_data.append({
            'curso': c,
            'dias_curso': dias_curso,
            'n_alumnos': len(alumnos),
            'perfectos': perfectos,
            'criticos': criticos,
            'sin_dias': (dias_curso == 0),
        })

    return render(request, 'alumnos/reporte_cursos_mes.html', {
        'mes': mes_str,
        'cursos_data': cursos_data,
    })
    
    
MESES_ES = {
    1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL",
    5: "MAYO", 6: "JUNIO", 7: "JULIO", 8: "AGOSTO",
    9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"
}

def mes_label(fecha):
    # fecha es date (día = 1)
    return f"{MESES_ES[fecha.month]} {fecha.year}"


@login_required
def exportar_excel(request):
    wb = Workbook()
    thin = Side(style="thin", color="CCCCCC")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="E9ECEF")
    title_font = Font(bold=True)
    encabezado_font = Font(bold=True)

    # Estilo porcentaje
    percent_style = NamedStyle(name="percent_style")
    percent_style.number_format = "0.00%"
    try:
        wb.add_named_style(percent_style)
    except ValueError:
        # Ya existe
        pass

    # Meses con cualquier dato
    meses_asistencia = AsistenciaMensual.objects.values_list("mes", flat=True)
    meses_dias = DiasClaseMensual.objects.values_list("mes", flat=True)
    meses = sorted(set(list(meses_asistencia) + list(meses_dias)))

    cursos = list(Curso.objects.all().order_by("nombre"))

    # Hoja de gráficos (activa)
    ws_graficos = wb.active
    ws_graficos.title = "Gráficos"

    # ========= Hojas por curso =========
    for curso in cursos:
        ws = wb.create_sheet(title=curso.nombre[:31])

        # Título
        end_col = 1 + max(0, (len(meses) * 4))
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col or 1)
        ws.cell(row=1, column=1, value=f"LISTA Y ASISTENCIA - {curso.nombre}").font = Font(bold=True, size=14)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

        # Encabezados
        ws.cell(row=2, column=1, value="ALUMNO").font = title_font
        ws.cell(row=2, column=1).fill = header_fill
        ws.cell(row=2, column=1).border = border_all
        ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

        col = 2
        for m in meses:
            etiqueta = mes_label(m)
            ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 3)
            h = ws.cell(row=2, column=col, value=etiqueta)
            h.font = title_font
            h.fill = header_fill
            h.alignment = Alignment(horizontal="center")
            h.border = border_all
            for j, txt in enumerate(["DÍAS CLASE", "PRESENTES", "INASISTENTES", "% ASISTENCIA"]):
                c = ws.cell(row=3, column=col + j, value=txt)
                c.font = title_font
                c.fill = header_fill
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                c.border = border_all
            col += 4

        # Datos por alumno
        alumnos = Alumno.objects.filter(curso=curso).order_by("nombre_completo")
        first_data_row = 4
        r = first_data_row
        for alumno in alumnos:
            ws.cell(row=r, column=1, value=alumno.nombre_completo).border = border_all
            ws.cell(row=r, column=1).alignment = Alignment(vertical="center")
            col = 2
            for m in meses:
                dias_obj = DiasClaseMensual.objects.filter(curso=curso, mes=m).first()
                dias = dias_obj.dias_clases if dias_obj else 0

                asis = AsistenciaMensual.objects.filter(alumno=alumno, mes=m).first()
                pres = asis.presentes if asis else 0
                inas = asis.inasistentes if asis else 0

                ws.cell(row=r, column=col,     value=dias).border = border_all
                ws.cell(row=r, column=col + 1, value=pres).border = border_all
                ws.cell(row=r, column=col + 2, value=inas).border = border_all

                # % = IF(dias>0, presentes/dias, 0)
                pct_cell = ws.cell(row=r, column=col + 3)
                dias_addr = ws.cell(row=r, column=col).coordinate
                pres_addr = ws.cell(row=r, column=col + 1).coordinate
                pct_cell.value = f"=IF({dias_addr}>0,{pres_addr}/{dias_addr},0)"
                pct_cell.style = "percent_style"
                pct_cell.border = border_all

                for j in range(4):
                    ws.cell(row=r, column=col + j).alignment = Alignment(horizontal="center")
                col += 4
            r += 1

        last_data_row = r - 1

        if first_data_row <= last_data_row:
            total_alumnos = len(alumnos)

            # Fila: Total alumnos
            resumen_row_1 = last_data_row + 2
            ws.cell(row=resumen_row_1, column=1, value="Total alumnos").font = title_font
            ws.cell(row=resumen_row_1, column=2, value=total_alumnos)

            # Fila: Resumen por mes (DÍAS CLASE y % ASISTENCIA DEL CURSO)
            resumen_row_2 = resumen_row_1 + 1
            ws.cell(row=resumen_row_2, column=1, value="Resumen por mes").font = title_font

            col = 2
            for m in meses:
                dias_obj = DiasClaseMensual.objects.filter(curso=curso, mes=m).first()
                dias = dias_obj.dias_clases if dias_obj else 0

                col_dias = col
                col_pres = col + 1
                col_pct  = col + 3

                # DÍAS CLASE
                ws.cell(row=resumen_row_2, column=col_dias, value=dias).border = border_all
                ws.cell(row=resumen_row_2, column=col_dias).alignment = Alignment(horizontal="center")

                # % asistencia del curso = SUM(presentes) / (dias * total_alumnos)
                pres_col_letter = get_column_letter(col_pres)
                sum_pres = f"=SUM({pres_col_letter}{first_data_row}:{pres_col_letter}{last_data_row})"
                dias_addr = ws.cell(row=resumen_row_2, column=col_dias).coordinate
                pct_cell = ws.cell(row=resumen_row_2, column=col_pct)
                if dias > 0 and total_alumnos > 0:
                    pct_cell.value = f"=IF({dias_addr}>0,({sum_pres})/({dias_addr}*{total_alumnos}),0)"
                else:
                    pct_cell.value = 0
                pct_cell.style = "percent_style"
                pct_cell.border = border_all
                pct_cell.alignment = Alignment(horizontal="center")

                col += 4

            # Anchos y congelar
            ws.column_dimensions['A'].width = 40
            for c in range(2, 2 + len(meses) * 4):
                ws.column_dimensions[get_column_letter(c)].width = 14
            ws.freeze_panes = "A4"
        else:
            ws.cell(row=4, column=1, value="(Sin alumnos en este curso)").font = Font(italic=True, color="888888")

    # ========= Hoja “Gráficos” (tabla y barras) =========
    ws = ws_graficos
    # Limpiar por si existía contenido anterior
    if ws.max_row > 1:
        for _ in range(ws.max_row):
            ws.delete_rows(1)

    # Último mes con datos
    mes_seleccionado = meses[-1] if meses else None
    etiqueta_mes = mes_label(mes_seleccionado) if mes_seleccionado else "SIN DATOS"

    ws.cell(row=1, column=1, value="Curso").font = encabezado_font
    ws.cell(row=1, column=2, value=f"% Asistencia {etiqueta_mes}").font = encabezado_font

    fila = 2
    for curso in cursos:
        pct = 0.0
        if mes_seleccionado:
            dias_obj = DiasClaseMensual.objects.filter(curso=curso, mes=mes_seleccionado).first()
            dias = dias_obj.dias_clases if dias_obj else 0
            if dias > 0:
                total_alumnos = Alumno.objects.filter(curso=curso).count()
                if total_alumnos > 0:
                    presentes_total = 0
                    for a in Alumno.objects.filter(curso=curso):
                        asis = AsistenciaMensual.objects.filter(alumno=a, mes=mes_seleccionado).first()
                        presentes_total += (asis.presentes if asis else 0)
                    pct = presentes_total / float(dias * total_alumnos)

        ws.cell(row=fila, column=1, value=curso.nombre)
        c = ws.cell(row=fila, column=2, value=pct)
        c.number_format = "0.00%"
        fila += 1

    # Configurar gráfico de barras (simple y funcional)
    num_cursos = len(cursos)
    if num_cursos > 0:
        chart = BarChart()
        chart.title = f"Asistencia por Curso - {etiqueta_mes}"
        chart.y_axis.title = "Porcentaje de asistencia"
        chart.y_axis.number_format = "0%"
        chart.x_axis.title = "Curso"
        chart.style = 2

        data = Reference(ws, min_col=2, min_row=1, max_row=num_cursos + 1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=num_cursos + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "D2")

    # Ajustes visuales tabla
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 22

    # ===== Responder archivo =====
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"asistencia_export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    resp = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp